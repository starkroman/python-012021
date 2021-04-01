import pandas as pd
import openpyxl

#hodiny = pd.read_excel('celkem.xlsx')
#print(hodiny.head())

#-------------------------------------

from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]

# Určím si barvu
sediva_barva = PatternFill("solid", fgColor="00F0F0F0")
zluta = PatternFill("solid", fgColor="00FFFF00")
orange = PatternFill("solid", fgColor="00FF6600")

radek = 1
for den in rozvrh_hodin:
  sloupec = 1
  for predmet in den:
    bunka = ws1.cell(radek, sloupec)
    bunka.value = predmet
    if radek % 2 == 1:
      bunka.fill = sediva_barva
    if predmet == "Matematika":
      bunka.fill = zluta
    if predmet == "Fyzika":
      bunka.fill = orange
    sloupec += 1
  radek += 1

wb.save(filename="rozvrh_hodin.xlsx")